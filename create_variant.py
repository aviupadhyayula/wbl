from google.cloud import dialogflowcx_v3
from openpyxl import Workbook, load_workbook
from utils import *
import asyncio
import time

# export GOOGLE_APPLICATION_CREDENTIALS=""

ROUTE_MAP = ""
ENDPOINT = "us-central1-dialogflow.googleapis.com"
VARIANT_NAME = ""

async def create_variant():
    agent = dialogflowcx_v3.Agent()
    agent.display_name = VARIANT_NAME
    agent.default_language_code = "en"
    agent.time_zone = "GMT-8:00"
    client = dialogflowcx_v3.AgentsAsyncClient(client_options={"api_endpoint": ENDPOINT})
    request = dialogflowcx_v3.CreateAgentRequest(parent="projects/heartschat-prod-a505/locations/us-central1", agent=agent)
    agent = await client.create_agent(request=request)
    return agent.name

async def export_reference():
    workbook = load_workbook(ROUTE_MAP)
    ref_sheet = workbook["Reference"]
    client = dialogflowcx_v3.AgentsAsyncClient(client_options={"api_endpoint": ENDPOINT})
    request = dialogflowcx_v3.ExportAgentRequest(name=ref_sheet["A1"].value)
    operation = await client.export_agent(request=request)
    response = await operation.result()
    return response.agent_content

async def restore_reference(agent_name, agent_content):
    client = dialogflowcx_v3.AgentsAsyncClient(client_options={"api_endpoint": ENDPOINT})
    request = dialogflowcx_v3.RestoreAgentRequest(name=agent_name, agent_content=agent_content)
    operation = await client.restore_agent(request=request)
    response = await operation.result()

async def update_variant(pages):
    workbook = load_workbook(ROUTE_MAP)
    ref_sheet = workbook["Reference"]
    async for page in pages:
        if page.transition_route_groups:
            for route_group_name in page.transition_route_groups:
                route_group = await get_route_group(route_group_name)
                for route in route_group.transition_routes:
                    route_num = await find_route_num(workbook, ref_sheet, route.name)
                    route_sheet = workbook[str(route_num)]
                    await update_route(route, route_sheet)
                await update_route_group(route_group)
        if page.transition_routes:
            for route in page.transition_routes:
                route_num = await find_route_num(workbook, ref_sheet, route.name)
                route_sheet = workbook[str(route_num)]
                await update_route(route, route_sheet)
        await update_page(page)

async def find_route_num(workbook, ref_sheet, route_name):
    for i in range(1, len(workbook.sheetnames)):
        if route_name == ref_sheet["B{}".format(i)].value:
            return i

async def update_route(route, route_sheet):
    for message in route.trigger_fulfillment.messages:
        fulfillment_num = 1
        for j in range(len(message.text.text)):
            if route_sheet["B{}".format(fulfillment_num)].value:
                message.text.text[j] = route_sheet["B{}".format(fulfillment_num)].value
                fulfillment_num += 1
            else:
                message.text.text = message.text.text[:j]
                break
        for j in range(fulfillment_num, 999):
            if route_sheet["C{}".format(j)].value:
                message.text.text.append(route_sheet["C{}".format(j)].value)
            else:
                break

async def main():
    variant_name = await create_variant()
    ref_content = await export_reference()
    await restore_reference(variant_name, ref_content)
    pages = await get_pages("{}/flows/00000000-0000-0000-0000-000000000000".format(variant_name))
    await update_variant(pages)

if __name__ == '__main__':
    asyncio.run(main())
