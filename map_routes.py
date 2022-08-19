from google.cloud import dialogflowcx_v3
from utils import *
from openpyxl import Workbook
from datetime import datetime
import asyncio
import string
import time

<<<<<<< HEAD
# export GOOGLE_APPLICATION_CREDENTIALS="heartschat-prod-a505-9599eda00cef.json"
# sassy = daa993a0-ac60-4bde-8b6b-6a42899b2329
# normal = 8b87e314-ec79-4386-9b5d-9347da0a0e23
=======
# export GOOGLE_APPLICATION_CREDENTIALS=""
>>>>>>> 9e6469b1682b90651f54084db882a19d64d05638

ALPHABET = list(string.ascii_uppercase)
ENDPOINT = "us-central1-dialogflow.googleapis.com"
AGENT = ""
FLOW = "{}/flows/00000000-0000-0000-0000-000000000000".format(AGENT)

route_num = 1
page_num = 0
fulfillment_num = 1
intent_num = 2

async def get_intent(intents, intent_name):
    async for intent in intents:
        if intent.name == intent_name:
            return intent
    client = dialogflowcx_v3.IntentsAsyncClient(client_options={"api_endpoint": ENDPOINT})
    request = dialogflowcx_v3.GetIntentRequest(name=intent_name)
    intent = await client.get_intent(request=request)
    return intent

def create_workbook():
    workbook = Workbook()
    routes_sheet = workbook.active
    routes_sheet.title = "Routes"
    fulfillments_sheet = workbook.create_sheet("Fulfillments")
    ref_sheet = workbook.create_sheet("Reference")
    ref_sheet["A1"] = AGENT
    ref_sheet["A2"] = FLOW
    ref_sheet["A3"] = ENDPOINT
    now = datetime.now()
    datetime_string = now.strftime("%m-%d-%Y_%H-%M-%S")
    return workbook, "route_map_{}.xlsx".format(datetime_string)  


async def write_route(workbook, route, intents):
    global route_num, fulfillment_num, page_num, intent_num
    ref_sheet = workbook["Reference"]
    fulfillments_sheet = workbook["Fulfillments"]
    routes_sheet = workbook["Routes"]
    route_sheet = workbook.create_sheet(str(route_num))
    ref_sheet["B{}".format(route_num)] = route.name
    routes_sheet["{}{}".format(ALPHABET[page_num], intent_num)].hyperlink = "#{}!A1".format(route_num)
    message_num = 1
    for message in route.trigger_fulfillment.messages:
        for text in message.text.text:
            route_sheet["B{}".format(message_num)] = text
            fulfillments_sheet["A{}".format(fulfillment_num)] = text
            message_num += 1
            fulfillment_num += 1
    if route.condition:
        routes_sheet["{}{}".format(ALPHABET[page_num], intent_num)] = route.condition
    if route.intent:
        intent = await get_intent(intents, route.intent)
        routes_sheet["{}{}".format(ALPHABET[page_num], intent_num)] = intent.display_name
        await write_intent(route_sheet, intent)

async def write_intent(route_sheet, intent):
    message_num = 1
    for phrase in intent.training_phrases:
        route_sheet["A{}".format(message_num)] = "".join([part.text for part in phrase.parts])
        message_num += 1

async def write_routes(pages, intents):
    global page_num, route_num, fulfillment_num, intent_num
    workbook, workbook_name = create_workbook()
    routes_sheet = workbook["Routes"]
    fulfillments_sheet = workbook["Fulfillments"]
    async for page in pages:
        routes_sheet["{}1".format(ALPHABET[page_num])] = page.display_name
        intent_num = 2
        if page.transition_route_groups:
            for route_group_name in page.transition_route_groups:
                route_group = await get_route_group(route_group_name)
                for route in route_group.transition_routes:
                    await write_route(workbook, route, intents)
                    intent_num += 1
                    route_num += 1
        if page.transition_routes:
            for route in page.transition_routes:
                await write_route(workbook, route, intents)
                intent_num += 1
                route_num += 1
        page_num += 1
    workbook.save(filename=workbook_name)

async def main():
    pages = await get_pages(FLOW)
    intents = await get_intents(AGENT)
    await write_routes(pages, intents)

if __name__ == '__main__':
    asyncio.run(main())
