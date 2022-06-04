from google.cloud import dialogflowcx_v3
from openpyxl import Workbook
import asyncio
import string
import re
import time
import json
from datetime import datetime

# export GOOGLE_APPLICATION_CREDENTIALS="heartschat-prod-a505-de929d994427.json"

ALPHABET = list(string.ascii_uppercase)
PROJECT_ID = "heartschat-prod-a505"
AGENT_ID = "68ce8bd9-e5ba-4119-8773-d050207969b7"
LOCATION_ID = "us-central1"
ENDPOINT_ID = "us-central1-dialogflow.googleapis.com"
FLOW_ID = "00000000-0000-0000-0000-000000000000"
AGENT = "projects/{}/locations/{}/agents/{}".format(PROJECT_ID, LOCATION_ID, AGENT_ID)
FLOW = "{}/flows/{}".format(AGENT, FLOW_ID)

count = 0
fulfillments = []

async def write_intents(intents):
    workbook = Workbook()
    sheet = workbook.active
    row = 1
    async for intent in intents:
        intent.display_name = intent.display_name.replace("?", "")
        sheet["A{}".format(row)] = intent.display_name
        intent_sheet = workbook.create_sheet(intent.display_name)
        sub_row = 1
        for phrase in intent.training_phrases:
            phrase_text = "".join([part.text for part in phrase.parts])
            intent_sheet["A{}".format(sub_row)] = phrase_text
            sub_row += 1
        row += 1
    workbook.save(filename="dialog_map.xlsx")

async def get_flows():
    client = dialogflowcx_v3.FlowsAsyncClient(client_options={"api_endpoint": ENDPOINT_ID})
    request = dialogflowcx_v3.GetFlowRequest(name=FLOW)
    flow = await client.get_flow(request=request)
    with open('flows.txt', 'w') as f:
        f.write(str(flow))

intent_count = 1

async def get_route_groups():
    client = dialogflowcx_v3.TransitionRouteGroupsAsyncClient(client_options={"api_endpoint": ENDPOINT_ID})
    request = dialogflowcx_v3.ListTransitionRouteGroupsRequest(parent=FLOW)
    route_groups = await client.list_transition_route_groups(request=request)
    await write_route_groups(route_groups)

async def testing():
    client = dialogflowcx_v3.IntentsAsyncClient(client_options={"api_endpoint": ENDPOINT_ID})
    request = dialogflowcx_v3.ListIntentsRequest(parent=AGENT)
    intents = await client.list_intents(request=request)
    with open('get_intents.txt', 'w') as f:
        async for intent in intents:
            f.write(str(intent.name))
            f.write("\n")
    client = dialogflowcx_v3.TransitionRouteGroupsAsyncClient(client_options={"api_endpoint": ENDPOINT_ID})
    request = dialogflowcx_v3.ListTransitionRouteGroupsRequest(parent=FLOW)
    route_groups = await client.list_transition_route_groups(request=request)
    with open('get_routes.txt', 'w') as f:
        async for route_group in route_groups:
            for route in route_group.transition_routes:
                f.write(str(route.intent))
                f.write("\n")

async def get_intents():
    client = dialogflowcx_v3.IntentsAsyncClient(client_options={"api_endpoint": ENDPOINT_ID})
    request = dialogflowcx_v3.ListIntentsRequest(parent=AGENT)
    intents = await client.list_intents(request=request)
    # await write_intents(intents)
    with open('intents.txt', 'w') as f:
        async for intent in intents:
            f.write("".join([str(ord(i)) for i in intent.name]))
    return intents

async def get_intent(intent_name, intents):
    # client = dialogflowcx_v3.IntentsAsyncClient(client_options={"api_endpoint": ENDPOINT_ID})
    # request = dialogflowcx_v3.GetIntentRequest(name=intent_id)
    # intent = await client.get_intent(request=request)
    # return intent
    async for intent in intents:
        if intent_name == intent.name:
            print("reg")
            return intent
    client = dialogflowcx_v3.IntentsAsyncClient(client_options={"api_endpoint": ENDPOINT_ID})
    request = dialogflowcx_v3.GetIntentRequest(name=intent_name)
    global count
    if count >= 55:
        print("timeout")
        time.sleep(60)
        count = 0
    intent = await client.get_intent(request = request)
    count += 1
    print("api")
    return intent

async def write_route_groups(route_groups):
    global intent_count
    workbook = Workbook()
    sheet = workbook.active
    col = 0
    intents = await get_intents()
    async for route_group in route_groups:
        sheet["{}1".format(ALPHABET[col])] = route_group.display_name
        intent_row = 2
        for route in route_group.transition_routes:
            intent = await get_intent(route.intent, intents)
            sheet["{}{}".format(ALPHABET[col], intent_row)] = intent.display_name
            intent_sheet = workbook.create_sheet(str(intent_count))
            sheet["{}{}".format(ALPHABET[col], intent_row)].hyperlink = "#{}!A1".format(str(intent_count))
            intent_count += 1
            phrase_row = 1
            for phrase in intent.training_phrases:
                phrase_text = "".join([part.text for part in phrase.parts])
                intent_sheet["A{}".format(phrase_row)] = phrase_text
                phrase_row += 1
            phrase_row = 1
            for message in route.trigger_fulfillment.messages:
                # json_obj = json.dumps(message.text)
                print(message)
                i = 0
                while True:
                    try:
                        intent_sheet["B{}".format(i+1)] = message.text.text[i]
                        print(message.text.text[i])
                        fulfillments.append(message.text.text[i])
                        i += 1
                        if i > 100:
                            break
                    except:
                        print("Num fulfillments found: ", str(i))
                        break
                    # phrase_row += 1
            intent_row += 1
        col += 1
    write_fulfillments(workbook)
    now = datetime.now()
    dt_string = now.strftime("%m-%d-%Y_%H-%M-%S")
    workbook.save(filename="dialog_map_{}.xlsx".format(dt_string))

def write_fulfillments(workbook):
    fulfillments_sheet = workbook.create_sheet("fulfillments")
    row = 1
    for f in fulfillments:
        fulfillments_sheet["A{}".format(row)] = f
        row += 1

def clean_string(s):
    rx = re.compile('\W+')
    s = rx.sub("", str(s)).strip()
    return s

def main():
    # asyncio.run(get_intents())
    asyncio.run(get_route_groups())

main()